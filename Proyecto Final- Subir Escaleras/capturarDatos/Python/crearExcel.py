from openpyxl import Workbook
from openpyxl import load_workbook

""" se crea un nuevo archivo de excel en la carpeta raiz """
wb=Workbook()   
wb.save("nuevo.xlsx")

wb2 = load_workbook('nuevo.xlsx')

ws1 = wb2.create_sheet('hoja1') #crear una hoja en excel al final
ws2 = wb2.create_sheet("hoja2",0) #crear una hoja en excel al inicio 

ws = wb2['hoja2']
#ws['E5'] = 350
valor = '13,54'
valor.split(sep=', ', maxsplit=10)
print(valor[0])
for num in range(2, 30, 1):
    ws['E'+str(num)] = int(valor[0]+valor[1])
    #print(num)

wb2.save('nuevo.xlsx') #siempre se pone para guardar los cambios



""" ws1 = wb.create_sheet("hoja1")
ws2 = wb.create_sheet("hoja2",0)
ws3 = wb.create_sheet("hoja3",-1)   """
