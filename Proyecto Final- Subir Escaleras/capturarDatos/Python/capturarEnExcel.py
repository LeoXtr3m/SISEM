import serial 
import time
from openpyxl import Workbook
from openpyxl import load_workbook

""" se crea un nuevo archivo de excel en la carpeta raiz """
#wb=Workbook()   
#wb.save("nuevo.xlsx")

wb2 = load_workbook('nuevo.xlsx')

#ws1 = wb2.create_sheet('hoja1') #crear una hoja en excel al final
#ws2 = wb2.create_sheet("hoja2",0) #crear una hoja en excel al inicio 

ws = wb2['hoja2']


serialArduino = serial.Serial("COM4", 9600)
time.sleep(1)
cont = 1
#print(cad)
print("***********************")
print("**Capturando datos...**")
print("***********************")
while cont<256:
    cad = serialArduino.readline().decode('ascii')

    valor = cad
    valor.split(sep=',', maxsplit=20) 
    if(cont>2): 
        ws['I'+str(cont)] = int(valor[1] + valor[2] + valor[3] + valor[4]+ valor[5])   
        ws['J'+str(cont)] = int(valor[7] + valor[8] + valor[9] + valor[10]+ valor[11])  
    cont+=1

print("**      LEEI TODO        **")
wb2.save('nuevo.xlsx') #siempre se pone para guardar los cambios 
print("***********************")
print("**       FIN         **")
print("***********************")